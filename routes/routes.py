# =============================================================================
# routes/routes.py
# =============================================================================

from flask import (Blueprint, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file)
from functools import wraps
from io import BytesIO
from database.db import (
    verifier_utilisateur, verifier_utilisateur_par_id,
    get_tous_articles, get_article_par_id, get_article_par_ref_constructeur,
    creer_article, modifier_article, modifier_stock, supprimer_article,
    get_historique, get_historique_article, get_statistiques, clear_historique,
    get_tous_utilisateurs, creer_utilisateur, get_emails_alertes,
    reset_password_utilisateur, supprimer_utilisateur as db_supprimer_utilisateur,
    exporter_inventaire_excel, importer_depuis_excel, COLONNES_EXCEL
)
from routes.qr_service import generer_qr_inventaire, generer_qr_bytes, generer_qr_article_bytes

main = Blueprint("main", __name__)


# =============================================================================
# DÉCORATEURS
# =============================================================================

def login_requis(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            flash("Veuillez vous connecter.", "warning")
            return redirect(url_for("main.login"))
        if not verifier_utilisateur_par_id(session["user_id"]):
            session.clear()
            flash("Votre compte a été supprimé.", "warning")
            return redirect(url_for("main.login"))
        return f(*args, **kwargs)
    return wrapper


def admin_requis(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("main.login"))
        if not verifier_utilisateur_par_id(session["user_id"]):
            session.clear()
            flash("Votre compte a été supprimé.", "warning")
            return redirect(url_for("main.login"))
        if session.get("role") != "admin":
            flash("Accès refusé : droits administrateur requis.", "danger")
            return redirect(url_for("main.dashboard"))
        return f(*args, **kwargs)
    return wrapper


def operateur_requis(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("main.login"))
        if not verifier_utilisateur_par_id(session["user_id"]):
            session.clear()
            flash("Votre compte a été supprimé.", "warning")
            return redirect(url_for("main.login"))
        if session.get("role") == "lecteur":
            flash("Accès refusé : droits opérateur requis.", "danger")
            return redirect(url_for("main.dashboard"))
        return f(*args, **kwargs)
    return wrapper


# =============================================================================
# AUTHENTIFICATION
# =============================================================================

@main.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("main.dashboard"))
    return redirect(url_for("main.login"))


@main.route("/login", methods=["GET", "POST"])
def login():
    if "user_id" in session:
        return redirect(url_for("main.dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        user = verifier_utilisateur(username, password)
        if user:
            session["user_id"]     = user["id"]
            session["username"]    = user["username"]
            session["nom_complet"] = user["nom_complet"]
            session["role"]        = user["role"]
            flash(f"Bienvenue, {user['nom_complet']} !", "success")
            return redirect(url_for("main.dashboard"))
        flash("Identifiants incorrects.", "danger")
    return render_template("login.html")


@main.route("/logout")
def logout():
    nom = session.get("nom_complet", "")
    session.clear()
    flash(f"À bientôt, {nom} !", "info")
    return redirect(url_for("main.login"))


# =============================================================================
# DASHBOARD
# =============================================================================

@main.route("/dashboard")
@login_requis
def dashboard():
    stats      = get_statistiques()
    articles   = get_tous_articles()
    historique = get_historique(limite=10)
    alertes    = [a for a in articles if a["stock_bas"] or a["quantite"] == 0]
    return render_template("dashboard.html", stats=stats, alertes=alertes, historique=historique)


# =============================================================================
# INVENTAIRE
# =============================================================================

@main.route("/inventaire")
@login_requis
def inventaire():
    recherche        = request.args.get("recherche", "").lower()
    filtre_statut    = request.args.get("statut", "")
    filtre_magasin   = request.args.get("magasin", "")
    filtre_fournisseur = request.args.get("fournisseur", "")

    articles = get_tous_articles()

    magasins    = sorted({a["magasin"]    for a in articles if a.get("magasin")})
    fournisseurs = sorted({a["fournisseur"] for a in articles if a.get("fournisseur")})

    if recherche:
        articles = [a for a in articles if
                    recherche in (a.get("nom") or "").lower() or
                    recherche in (a.get("ref_constructeur") or "").lower() or
                    recherche in (a.get("ref_fournisseur") or "").lower() or
                    recherche in (a.get("description") or "").lower() or
                    recherche in (a.get("emplacement") or "").lower() or
                    recherche in (a.get("magasin") or "").lower() or
                    recherche in (a.get("fournisseur") or "").lower()]
    if filtre_magasin:
        articles = [a for a in articles if a.get("magasin") == filtre_magasin]
    if filtre_fournisseur:
        articles = [a for a in articles if a.get("fournisseur") == filtre_fournisseur]
    if filtre_statut == "bas":
        articles = [a for a in articles if a["stock_bas"] and a["quantite"] > 0]
    elif filtre_statut == "rupture":
        articles = [a for a in articles if a["quantite"] == 0]

    return render_template("inventaire.html",
        articles=articles, magasins=magasins, fournisseurs=fournisseurs,
        recherche=recherche, filtre_statut=filtre_statut,
        filtre_magasin=filtre_magasin, filtre_fournisseur=filtre_fournisseur)


@main.route("/inventaire/ajouter", methods=["GET", "POST"])
@operateur_requis
def ajouter_article():
    if request.method == "POST":
        ref_constructeur = request.form.get("ref_constructeur", "").strip().upper()
        ref_fournisseur  = request.form.get("ref_fournisseur", "").strip()
        nom              = request.form.get("nom", "").strip()
        description      = request.form.get("description", "").strip()
        quantite         = float(request.form.get("quantite", 0) or 0)
        quantite_min     = float(request.form.get("quantite_min", 5) or 5)
        unite            = request.form.get("unite", "unité").strip()
        emplacement      = request.form.get("emplacement", "").strip()
        magasin          = request.form.get("magasin", "").strip()
        fournisseur      = request.form.get("fournisseur", "").strip()

        if not ref_constructeur or not ref_fournisseur or not nom:
            flash("Réf. constructeur, réf. fournisseur et nom sont obligatoires.", "warning")
            return render_template("article_form.html", mode="ajouter")

        article_id = creer_article(ref_constructeur, ref_fournisseur, nom, description,
                                   quantite, quantite_min, unite, emplacement, magasin, fournisseur)
        if article_id:
            flash(f"Article '{nom}' créé avec succès !", "success")
            return redirect(url_for("main.inventaire"))
        flash(f"La référence constructeur '{ref_constructeur}' existe déjà.", "danger")

    return render_template("article_form.html", mode="ajouter")


@main.route("/inventaire/modifier/<int:article_id>", methods=["GET", "POST"])
@operateur_requis
def modifier_article_route(article_id):
    article = get_article_par_id(article_id)
    if not article:
        flash("Article introuvable.", "danger")
        return redirect(url_for("main.inventaire"))

    if request.method == "POST":
        action = request.form.get("action")

        if action == "infos":
            modifier_article(
                article_id,
                ref_fournisseur = request.form.get("ref_fournisseur", "").strip(),
                nom             = request.form.get("nom", "").strip(),
                description     = request.form.get("description", "").strip(),
                quantite_min    = float(request.form.get("quantite_min", 5) or 5),
                unite           = request.form.get("unite", "unité").strip(),
                emplacement     = request.form.get("emplacement", "").strip(),
                magasin         = request.form.get("magasin", "").strip(),
                fournisseur     = request.form.get("fournisseur", "").strip()
            )
            flash("Informations mises à jour.", "success")

        elif action == "stock":
            nouvelle_quantite = float(request.form.get("quantite", 0) or 0)
            type_mouvement    = request.form.get("type_mouvement", "correction")
            commentaire       = request.form.get("commentaire", "").strip()
            ok, sous_seuil, art = modifier_stock(
                article_id, nouvelle_quantite, type_mouvement,
                session["user_id"], commentaire,
                operateur_nom=session.get("nom_complet")
            )
            if ok and sous_seuil:
                _envoyer_alerte(art, nouvelle_quantite, operateur=session.get("nom_complet", ""))
            flash("Stock mis à jour.", "success")

        return redirect(url_for("main.modifier_article_route", article_id=article_id))

    return render_template("article_form.html", article=article, mode="modifier")


@main.route("/inventaire/supprimer/<int:article_id>", methods=["POST"])
@admin_requis
def supprimer_article_route(article_id):
    article = get_article_par_id(article_id)
    if article:
        supprimer_article(article_id)
        flash(f"Article '{article['nom']}' supprimé.", "success")
    else:
        flash("Article introuvable.", "danger")
    return redirect(url_for("main.inventaire"))


# =============================================================================
# IMPORT / EXPORT EXCEL
# =============================================================================

@main.route("/inventaire/exporter")
@login_requis
def exporter_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash("Le module openpyxl n'est pas installé.", "danger")
        return redirect(url_for("main.inventaire"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire SDEL"

    entetes = {
        "ref_constructeur": "Réf. Constructeur",
        "ref_fournisseur":  "Réf. Fournisseur",
        "nom":              "Nom",
        "description":      "Description",
        "quantite":         "Quantité",
        "quantite_min":     "Seuil min.",
        "unite":            "Unité",
        "emplacement":      "Emplacement",
        "magasin":          "Magasin",
        "fournisseur":      "Fournisseur",
    }

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(bold=True, color="F5A623")

    for col_idx, (key, label) in enumerate(entetes.items(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    donnees = exporter_inventaire_excel()
    for row_idx, ligne in enumerate(donnees, start=2):
        for col_idx, key in enumerate(entetes.keys(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=ligne.get(key, ""))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="inventaire_sdel.xlsx")


@main.route("/inventaire/importer", methods=["GET", "POST"])
@operateur_requis
def importer_excel():
    if request.method == "POST":
        fichier = request.files.get("fichier")
        if not fichier or not fichier.filename.endswith(".xlsx"):
            flash("Veuillez uploader un fichier .xlsx valide.", "warning")
            return redirect(url_for("main.importer_excel"))

        try:
            import openpyxl
        except ImportError:
            flash("Le module openpyxl n'est pas installé.", "danger")
            return redirect(url_for("main.inventaire"))

        try:
            wb = openpyxl.load_workbook(fichier)
            ws = wb.active

            entetes = [str(cell.value or "").strip().lower().replace(" ", "_").replace(".", "")
                       for cell in ws[1]]

            mapping = {
                "réf_constructeur": "ref_constructeur",
                "réf_fournisseur":  "ref_fournisseur",
                "ref_constructeur": "ref_constructeur",
                "ref_fournisseur":  "ref_fournisseur",
                "nom":              "nom",
                "description":      "description",
                "quantité":         "quantite",
                "quantite":         "quantite",
                "seuil_min":        "quantite_min",
                "quantite_min":     "quantite_min",
                "unité":            "unite",
                "unite":            "unite",
                "emplacement":      "emplacement",
                "magasin":          "magasin",
                "fournisseur":      "fournisseur",
            }

            lignes = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                ligne = {}
                for i, val in enumerate(row):
                    if i < len(entetes):
                        cle = mapping.get(entetes[i], entetes[i])
                        ligne[cle] = val
                lignes.append(ligne)

            nb_crees, nb_mis_a_jour, erreurs = importer_depuis_excel(
                lignes, session["user_id"], operateur_nom=session.get("nom_complet", "Import")
            )

            if nb_crees or nb_mis_a_jour:
                flash(f"{nb_crees} article(s) créé(s), {nb_mis_a_jour} stock(s) mis à jour.", "success")
            for err in erreurs:
                flash(err, "warning")

        except Exception as e:
            flash(f"Erreur lors de la lecture du fichier : {e}", "danger")

        return redirect(url_for("main.inventaire"))

    return render_template("import_excel.html")


# =============================================================================
# SCAN QR CODE — public sans login
# =============================================================================

@main.route("/scan/article/<int:article_id>", methods=["GET"])
def scan_article(article_id):
    article = get_article_par_id(article_id)
    if not article:
        return "Article introuvable", 404
    historique = get_historique_article(article_id, limite=10)
    return render_template("scan_article.html", article=article, historique=historique)


@main.route("/scan/article/<int:article_id>/mouvement", methods=["POST"])
def scan_mouvement(article_id):
    article = get_article_par_id(article_id)
    if not article:
        return "Article introuvable", 404

    type_mouvement  = request.form.get("type_mouvement", "entree")
    quantite_saisie = float(request.form.get("quantite", 0) or 0)
    operateur_nom   = request.form.get("operateur_nom", "").strip()
    emplacement     = request.form.get("emplacement", "").strip()
    commentaire     = request.form.get("commentaire", "").strip()

    if not operateur_nom:
        flash("Veuillez indiquer votre nom.", "warning")
        return redirect(url_for("main.scan_article", article_id=article_id))

    quantite_actuelle = float(article["quantite"])
    if type_mouvement == "entree":
        nouvelle_quantite = quantite_actuelle + quantite_saisie
    elif type_mouvement == "sortie":
        nouvelle_quantite = max(0, quantite_actuelle - quantite_saisie)
    else:
        nouvelle_quantite = quantite_saisie

    ok, sous_seuil, art = modifier_stock(
        article_id, nouvelle_quantite, type_mouvement,
        user_id=None, commentaire=commentaire,
        operateur_nom=operateur_nom, emplacement=emplacement or None
    )
    if ok and sous_seuil:
        _envoyer_alerte(art, nouvelle_quantite, operateur=operateur_nom, emplacement=emplacement)

    flash("Mouvement enregistré ✓", "success")
    return redirect(url_for("main.scan_article", article_id=article_id))


def _envoyer_alerte(article_row, nouvelle_quantite, operateur="Système", emplacement=""):
    try:
        from routes.mail_service import envoyer_alerte_stock
        destinataires = get_emails_alertes()
        if not destinataires:
            return
        envoyer_alerte_stock(
            article_nom   = article_row.get("nom", ""),
            article_ref   = article_row.get("ref_constructeur", ""),
            quantite      = nouvelle_quantite,
            quantite_min  = article_row.get("quantite_min", 0),
            emplacement   = emplacement or article_row.get("emplacement", ""),
            magasin       = article_row.get("magasin", ""),
            operateur     = operateur,
            destinataires = destinataires
        )
    except Exception as e:
        print(f"[MAIL] Erreur alerte : {e}")


# =============================================================================
# HISTORIQUE
# =============================================================================

@main.route("/historique")
@login_requis
def historique():
    limite = int(request.args.get("limite", 50))
    data   = get_historique(limite=limite)
    return render_template("historique.html", historique=data, limite=limite)


@main.route("/historique/clear", methods=["POST"])
@admin_requis
def clear_historique_route():
    clear_historique()
    flash("Historique effacé.", "success")
    return redirect(url_for("main.historique"))


# =============================================================================
# QR CODES
# =============================================================================

@main.route("/qrcode")
@login_requis
def qrcode_page():
    articles = get_tous_articles()
    return render_template("qrcode.html", articles=articles)


@main.route("/qrcode.png")
def qrcode_image():
    data = generer_qr_bytes()
    return send_file(BytesIO(data), mimetype="image/png", download_name="qrcode_inventaire.png")


@main.route("/qrcode/article/<int:article_id>.png")
def qrcode_article_image(article_id):
    article = get_article_par_id(article_id)
    if not article:
        return "Article introuvable", 404
    data = generer_qr_article_bytes(article_id, article["nom"], article["ref_constructeur"])
    return send_file(BytesIO(data), mimetype="image/png",
                     download_name=f"qr_{article['ref_constructeur']}.png")


@main.route("/qrcode/regenerer", methods=["POST"])
@operateur_requis
def regenerer_qrcode():
    generer_qr_inventaire()
    flash("QR Code régénéré.", "success")
    return redirect(url_for("main.qrcode_page"))


# =============================================================================
# PAGE PUBLIQUE
# =============================================================================

@main.route("/public/inventaire")
def inventaire_public():
    recherche = request.args.get("q", "").strip().lower()
    articles  = get_tous_articles()
    stats     = get_statistiques()
    if recherche:
        articles = [a for a in articles if
                    recherche in (a.get("nom") or "").lower() or
                    recherche in (a.get("ref_constructeur") or "").lower() or
                    recherche in (a.get("ref_fournisseur") or "").lower() or
                    recherche in (a.get("emplacement") or "").lower() or
                    recherche in (a.get("magasin") or "").lower()]
    return render_template("inventaire_public.html",
        articles=articles, stats=stats, recherche=recherche,
        total_articles=len(get_tous_articles()))


# =============================================================================
# UTILISATEURS
# =============================================================================

@main.route("/utilisateurs")
@admin_requis
def utilisateurs():
    users = get_tous_utilisateurs()
    return render_template("utilisateurs.html", users=users)


@main.route("/utilisateurs/ajouter", methods=["POST"])
@admin_requis
def ajouter_utilisateur():
    username    = request.form.get("username", "").strip()
    password    = request.form.get("password", "").strip()
    nom_complet = request.form.get("nom_complet", "").strip()
    role        = request.form.get("role", "lecteur")
    if not username or not password:
        flash("Identifiant et mot de passe obligatoires.", "warning")
        return redirect(url_for("main.utilisateurs"))
    if creer_utilisateur(username, password, nom_complet, role):
        flash(f"Utilisateur '{username}' créé.", "success")
    else:
        flash(f"L'identifiant '{username}' est déjà pris.", "danger")
    return redirect(url_for("main.utilisateurs"))


@main.route("/utilisateurs/reset-password/<int:user_id>", methods=["POST"])
@admin_requis
def reset_password_route(user_id):
    nouveau_mdp = request.form.get("nouveau_mdp", "").strip()
    if not nouveau_mdp:
        flash("Le mot de passe ne peut pas être vide.", "warning")
        return redirect(url_for("main.utilisateurs"))
    reset_password_utilisateur(user_id, nouveau_mdp)
    flash("Mot de passe réinitialisé.", "success")
    return redirect(url_for("main.utilisateurs"))


@main.route("/utilisateurs/supprimer/<int:user_id>", methods=["POST"])
@admin_requis
def supprimer_utilisateur(user_id):
    if user_id == session.get("user_id"):
        flash("Vous ne pouvez pas supprimer votre propre compte.", "danger")
        return redirect(url_for("main.utilisateurs"))
    db_supprimer_utilisateur(user_id)
    flash("Utilisateur supprimé.", "success")
    return redirect(url_for("main.utilisateurs"))


# =============================================================================
# API JSON
# =============================================================================

@main.route("/api/articles")
@login_requis
def api_articles():
    return jsonify(get_tous_articles())


@main.route("/api/stats")
@login_requis
def api_stats():
    return jsonify(get_statistiques())
